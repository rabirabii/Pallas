import json

from openpyxl import load_workbook

from pipeline.analytics import build_area_dataset
from pipeline.export import write_all_exports, write_csv, write_excel, write_json
from pipeline.speedhome.models import DailyRentAvailability, Furnishing, PropertyListing


def make_listing():
    return PropertyListing(
        id="abc123",
        title="Example",
        propertyName="Example Property",
        areaName="Mont Kiara",
        areaSlug="mont-kiara",
        segment="2BR",
        bedroomCount=2,
        bathroomCount=2,
        parkingCount=1,
        monthlyRentRM=3000,
        annualRentRM=36000,
        annualRentIsDerived=True,
        dailyRentRM=None,
        dailyRentAvailability=DailyRentAvailability.UNAVAILABLE,
        sizeSqft=1000,
        furnishing=Furnishing.FULLY_FURNISHED,
        minimumTenureMonths=12,
        moveInStatus=None,
        moveInDate="May 4, 2026",
        verified=True,
        zeroDeposit=True,
        sourceUrl="https://speedhome.com/details/example",
        sourcePage=1,
        scrapedAt="2026-06-23T00:00:00+00:00",
        parseWarnings=[],
        dataQualityFlags=[],
    )


def make_dataset():
    return build_area_dataset(
        area_name="Mont Kiara",
        area_slug="mont-kiara",
        source_url="https://speedhome.com/rent/mont-kiara",
        scraped_at="2026-06-23T00:00:00+00:00",
        listings=[make_listing()],
    )


def test_write_json(tmp_path):
    output_path = tmp_path / "mont-kiara.json"

    write_json(make_dataset(), output_path)

    data = json.loads(output_path.read_text(encoding="utf-8"))
    assert data["metadata"]["areaSlug"] == "mont-kiara"
    assert data["listings"][0]["monthlyRentRM"] == 3000


def test_write_csv(tmp_path):
    output_path = tmp_path / "mont-kiara.csv"

    write_csv(make_dataset(), output_path)

    text = output_path.read_text(encoding="utf-8")
    assert "Listing Title" in text
    assert "Example" in text


def test_write_excel(tmp_path):
    output_path = tmp_path / "mont-kiara.xlsx"

    write_excel(make_dataset(), output_path)

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "Listings",
        "Price Summary",
        "Data Quality",
        "Metadata",
        "Methodology",
    ]
    assert workbook["Listings"]["A1"].value == "ID"
    assert workbook["Listings"]["I2"].value == 3000


def test_write_all_exports(tmp_path):
    json_path, csv_path, excel_path = write_all_exports(
        make_dataset(),
        tmp_path / "area.json",
        tmp_path / "area.csv",
        tmp_path / "area.xlsx",
    )

    assert json_path.exists()
    assert csv_path.exists()
    assert excel_path.exists()