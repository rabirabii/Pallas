from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from pipeline.speedhome.models import AreaDataset, PropertyListing


LISTING_COLUMNS = [
    ("id", "ID"),
    ("title", "Listing Title"),
    ("propertyName", "Property Name"),
    ("areaName", "Area"),
    ("segment", "Unit Segment"),
    ("bedroomCount", "Bedrooms"),
    ("bathroomCount", "Bathrooms"),
    ("parkingCount", "Parking"),
    ("monthlyRentRM", "Monthly Rent (RM)"),
    ("annualRentRM", "Annual Rent (RM)"),
    ("annualRentIsDerived", "Annual Rent Derived"),
    ("dailyRentRM", "Daily Rent (RM)"),
    ("dailyRentAvailability", "Daily Rent Availability"),
    ("sizeSqft", "Size (sqft)"),
    ("furnishing", "Furnishing"),
    ("minimumTenureMonths", "Minimum Tenure (months)"),
    ("moveInStatus", "Move-in Status"),
    ("moveInDate", "Move-in Date"),
    ("verified", "Verified"),
    ("zeroDeposit", "Zero Deposit"),
    ("sourceUrl", "Source URL"),
    ("sourcePage", "Source Page"),
    ("scrapedAt", "Scraped At"),
    ("parseWarnings", "Parse Warnings"),
    ("dataQualityFlags", "Data Quality Flags"),
]

SUMMARY_COLUMNS = [
    ("segment", "Unit Segment"),
    ("unitCount", "Unit Count"),
    ("validPriceCount", "Valid Price Count"),
    ("averageMonthlyRentRM", "Average Monthly Rent (RM)"),
    ("medianMonthlyRentRM", "Median Monthly Rent (RM)"),
    ("modeMonthlyRentRM", "Mode Monthly Rent (RM)"),
    ("modeStatus", "Mode Status"),
    ("multipleModes", "Multiple Modes"),
    ("allModes", "All Modes"),
    ("fairPriceRM", "Fair Price (RM)"),
    ("fairPriceStatus", "Fair Price Status"),
    ("averageSizeSqft", "Average Size (sqft)"),
    ("dataConfidence", "Data Confidence"),
]


def _model_dump(model: Any) -> dict[str, Any]:
    return model.model_dump(mode="json")


def _cell_value(value: Any) -> Any:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return value


def listing_to_row(listing: PropertyListing) -> dict[str, Any]:
    dumped = _model_dump(listing)
    return {
        header: _cell_value(dumped.get(field))
        for field, header in LISTING_COLUMNS
    }


def summary_to_row(summary: Any) -> dict[str, Any]:
    dumped = _model_dump(summary)
    return {
        header: _cell_value(dumped.get(field))
        for field, header in SUMMARY_COLUMNS
    }


def write_json(dataset: AreaDataset, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    output_path.write_text(
        json.dumps(dataset.model_dump(mode="json"), indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    return output_path


def write_csv(dataset: AreaDataset, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = [listing_to_row(listing) for listing in dataset.listings]

    with output_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[header for _, header in LISTING_COLUMNS],
        )
        writer.writeheader()
        writer.writerows(rows)

    return output_path


def _append_dict_rows(worksheet, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)

    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def _apply_currency_format(worksheet, header_names: set[str]) -> None:
    if worksheet.max_row < 2:
        return

    headers = [cell.value for cell in worksheet[1]]

    for index, header in enumerate(headers, start=1):
        if header not in header_names:
            continue

        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=index).number_format = '"RM" #,##0'


def write_excel(dataset: AreaDataset, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()

    listings_sheet = workbook.active
    listings_sheet.title = "Listings"
    _append_dict_rows(
        listings_sheet,
        [listing_to_row(listing) for listing in dataset.listings],
    )
    _apply_currency_format(
        listings_sheet,
        {"Monthly Rent (RM)", "Annual Rent (RM)", "Daily Rent (RM)"},
    )

    summary_sheet = workbook.create_sheet("Price Summary")
    _append_dict_rows(
        summary_sheet,
        [summary_to_row(summary) for summary in dataset.summaries],
    )
    _apply_currency_format(
        summary_sheet,
        {
            "Average Monthly Rent (RM)",
            "Median Monthly Rent (RM)",
            "Mode Monthly Rent (RM)",
            "Fair Price (RM)",
        },
    )

    quality_sheet = workbook.create_sheet("Data Quality")
    _append_dict_rows(
        quality_sheet,
        [
            {
                "Metric": "Missing Price Count",
                "Value": dataset.qualityReport.missingPriceCount,
            },
            {
                "Metric": "Missing Size Count",
                "Value": dataset.qualityReport.missingSizeCount,
            },
            {
                "Metric": "Unknown Furnishing Count",
                "Value": dataset.qualityReport.unknownFurnishingCount,
            },
            {
                "Metric": "Warning Count",
                "Value": dataset.qualityReport.warningCount,
            },
        ],
    )

    metadata_sheet = workbook.create_sheet("Metadata")
    _append_dict_rows(
        metadata_sheet,
        [
            {"Field": key, "Value": _cell_value(value)}
            for key, value in dataset.metadata.model_dump(mode="json").items()
        ],
    )

    methodology_sheet = workbook.create_sheet("Methodology")
    _append_dict_rows(
        methodology_sheet,
        [
            {
                "Topic": "Snapshot architecture",
                "Description": "Data is scraped locally and exported as static files.",
            },
            {
                "Topic": "Monthly rent",
                "Description": "Monthly rent is the observed SPEEDHOME listing price when available.",
            },
            {
                "Topic": "Annual rent",
                "Description": "Annual rent is derived as monthlyRentRM * 12.",
            },
            {
                "Topic": "Daily rent",
                "Description": "Daily rent is left unavailable unless explicitly present in source data.",
            },
            {
                "Topic": "Fair Price",
                "Description": "Fair Price uses median for 3 samples, otherwise IQR outlier filtering plus median and trimmed mean.",
            },
        ],
    )

    workbook.save(output_path)
    return output_path


def write_all_exports(
    dataset: AreaDataset,
    json_path: Path,
    csv_path: Path,
    excel_path: Path,
) -> tuple[Path, Path, Path]:
    return (
        write_json(dataset, json_path),
        write_csv(dataset, csv_path),
        write_excel(dataset, excel_path),
    )